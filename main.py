#!/usr/bin/env python3
import collections
from typing import Dict, List
from collections import defaultdict
from openpyxl.worksheet.worksheet import Worksheet
import sys
from requests_html import HTMLSession
from bs4 import BeautifulSoup
import urllib3
from selenium import webdriver
import requests
urllib3.disable_warnings()

import openpyxl

URLS = [
    'https://meteoinfo.ru/hdmapsmeteoalarmszfo',
    'https://meteoinfo.ru/hdmapsmeteoalarmcfo',
    'https://meteoinfo.ru/hdmapsmeteoalarmprfo',
    'https://meteoinfo.ru/hdmapsmeteoalarmyugfo',
    'https://meteoinfo.ru/hdmapsmeteoalarmyskfo',
    'https://meteoinfo.ru/hdmapsmeteoalarmural',
    'https://meteoinfo.ru/hdmapsmeteoalarmsyb',
    'https://meteoinfo.ru/hdmapsmeteoalarmdv',
]

HAZARDS_URL = 'https://meteoinfo.ru/hazardsbull'

SIZE = 4

SHEET_NAME = 'Билет по регионам'

REGION_COLUMN = 2

FORECAST_COLUMNS = (6, 7, 8, 9)
FACTOR_COLUMNS = (10, 11, 12, 13)

FORCASTS = (
    'Угроза повреждения ЛЭП и линий связи, обрушения слабо укрепленных, широкоформатных, ветхих и рекламных конструкций',
    'Угроза выхода из строя систем ЖКХ, городских коммуникаций',
    'Угроза затопления населенных пунктов, дорог, мостов',
    'Угроза сбоев в работе всех видов транспорта, увеличение количества ДТП',
    'Угроза жизни и здоровью населения',
)

INTERCHANGEABLE_FACTORS = [
    {'Ветер', 'Дождь', 'Снег/Обледенение', 'Гроза'},
    {'Заморозки', 'Очень низкая температура', 'Очень высокая температура'},
    {'Туман', 'Голеледно - изморозевые отложения'},
    {'Пожарная опасность'},
    {'Паводок'}
]

FORCAST_MAP = {
    'Ветер': set((FORCASTS[0], FORCASTS[1])),
    'Заморозки': set((FORCASTS[4],)),
    'Туман': set((FORCASTS[3],)),
    'Очень низкая температура': set((FORCASTS[4],)),
    'Пожарная опасность': set((FORCASTS[2],)),
    'Дождь': set((FORCASTS[0], FORCASTS[1])),
    'Паводок': set((FORCASTS[2],)),
    'Пыльная (песчаная) буря': set((FORCASTS[0], FORCASTS[1], FORCASTS[4])),
    # 'прочие опасности': '',
    'Снег/Обледенение': set((FORCASTS[0], FORCASTS[1])),
    'Гроза': set((FORCASTS[0], FORCASTS[1])),
    'Очень высокая температура': set((FORCASTS[4],)),
    # 'прибрежные события': '',
    # 'лавины': '',
    # 'наводнения': '',
    # 'сель': '',
    'Голеледно - изморозевые отложения': set((FORCASTS[3],)),
}

FACTORS = FORCAST_MAP.keys()

regions_need_to_join = {'саха',
                        'магаданская',
                        'ямало ненецкий ао',
                        'камчатский край',
                        'амурская',
                        'хабаровский край',
                        'чеч',
                        'северная осетия'}


def join_region(region: str) -> str:
    if region.startswith('сахалин'):
        return 'сахалин'
    for reg in regions_need_to_join:
        if region.startswith(reg):
            return reg
    return region


def normolize_string(string: str):
    string = string.strip().lower().translate({ord('.'): None})
    string = ' '.join(string.split('-'))

    # remove stop words
    stop_words = ['г', 'обл', 'область', 'республика']
    try:
        splitted_string = string.split()
    except IndexError:
        return ''
    string = ' '.join(w for w in splitted_string if w not in stop_words)
    return join_region(string)


def get_regions(urls: List[str]) -> Dict:
    regions = defaultdict(list)
    session = HTMLSession()




    for url in urls:
        # content = requests.get(url, verify=False).content
        #
        # driver = webdriver.Chrome(executable_path='/Users/dinar/internship/parser_mchs/chromedriver')

        options = webdriver.ChromeOptions()
        options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
        chrome_driver_binary = "/usr/local/bin/chromedriver"
        driver = webdriver.Chrome(chrome_driver_binary, chrome_options=options)
        # driver.get("data:text/html;charset=utf-8," + dri)

        #
        # driver = webdriver.Firefox(executable_path='/Users/dinar/internship/parser_mchs/geckodriver')
        # driver.get(url)
        # p_element = driver.find_element_by_id(id_='intro-text')
        # print(p_element.text)

        # response = session.get(url, verify=False)
        # response.html.render()
        # rows = response.html.find('#tbl_1', first=True).xpath('*/table/tbody/tr')
        # for row in rows:
        #     row_data = row.xpath('//div[@class="test"]/@data-original-title', first=True)
        #     try:
        #         row_data_bs = BeautifulSoup(row_data, features='lxml')
        #     except:
        #         continue  # if raw_data is None
        #     region_name = normolize_string(row_data_bs.find(class_='cl_tbl_obl_name_tooltip').text)
        #     region_factor = row_data_bs.find(class_='cl_tbl_tooltip_text').find('b').text.strip('. ')
        #
        #     if region_factor not in regions[region_name] and region_factor != '':
        #         regions[region_name].append(region_factor)

    return regions


def remove_interchangeable_factors(factors: list):
    factors_set = set()
    prev_len_factors_set = 0

    for factor in factors:
        forecast = FORCAST_MAP.get(factor)
        if not forecast:
            continue
        factors_set.update(forecast)
        if len(factors_set) == prev_len_factors_set:
            factors.remove(factor)
            return factors
        prev_len_factors_set = len(factors_set)
    return factors


def analyze_factors(factors: list):
    final_forecasts = set()

    prev_len_factors = len(factors)
    while len(factors) > SIZE:
        factors = remove_interchangeable_factors(factors)
        if len(factors) == prev_len_factors:  # If nothing has been deleted, means there is no interchangeable factors
            factors = factors[:SIZE]
        prev_len_factors = len(factors)

    for factor in factors:
        forecast: set = FORCAST_MAP.get(factor)
        if not forecast:
            continue
        final_forecasts.update(forecast)

    while len(final_forecasts) > SIZE:
        final_forecasts.pop()

    return final_forecasts, set(factors)


def fill_excel(ws: Worksheet, regions: dict):
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=REGION_COLUMN).value
        if not name:
            continue

        name = normolize_string(name)
        factors = regions.get(name)

        if not factors:
            continue

        forecasts, factors = analyze_factors(factors)
        for column, forecast in zip(FORECAST_COLUMNS, forecasts):
            ws.cell(row=row, column=column).value = forecast

        for column, factor in zip(FACTOR_COLUMNS, factors):
            ws.cell(row=row, column=column).value = factor


def get_hazards(url: str):
    session = HTMLSession(browser_args=["--proxy-server=201.249.161.51:999"])
    response = session.get(url, verify=False)
    rows = response.html.find('#div_1', first=True).xpath('*/table/tbody/tr')
    row_data = rows[1].xpath('//p', first=True)
    print(row_data.text)

def main(filename: str):
    # wb = openpyxl.load_workbook(filename)
    # ws = wb[SHEET_NAME]

    regions = get_regions(urls=URLS)
    # fill_excel(ws=ws, regions=regions)
    #
    # wb.save(filename)
    # get_hazards(HAZARDS_URL)


if __name__ == '__main__':
    with open('conf', 'r') as f:
        filename = f.read().strip()
    main(filename)
