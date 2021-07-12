import openpyxl
from typing import Dict, List, Tuple

from bs4 import BeautifulSoup
from requests_html import HTMLSession
import urllib3
from openpyxl.worksheet.worksheet import Worksheet
from nltk.stem.snowball import SnowballStemmer
from datetime import datetime, date, time, timedelta
from openpyxl.formula import Tokenizer
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import get_column_letter
import dateparser
from copy import copy
from openpyxl.formatting.rule import FormulaRule

import re
from openpyxl.styles import Color, PatternFill, Font, Border

urllib3.disable_warnings()


URL = 'https://yastat.net/s3/milab/2020/covid19-stat/data/v10/deep_data.json'
REGION_COLUMN = 2
REGION_ROWS_RANGE = (2, 87)
WEEKLY_REPORT_DAY = 5
DAYS_OF_THE_WEEK_MAP = {
    1: 'ПН',
    2: 'ВТ',
    3: 'СР',
    4: 'ЧТ',
    5: 'ПТ',
    6: 'СБ',
    7: 'ВС',
}

def normolize_string(string: str):
    string = string.strip().lower().translate({ord('.'): None})
    string = ' '.join(string.split('-'))

    # remove stop words
    stop_words = ['г', 'обл', 'область', 'республика', 'автономная', 'ао']
    try:
        splitted_string = string.split()
    except IndexError:
        return ''
    stemmer = SnowballStemmer("russian")
    string = ' '.join(stemmer.stem(w) for w in splitted_string if w not in stop_words)
    return string


def fill_сases_sheet(ws: Worksheet, info: dict):
    OVERALL_CASES_FORMULA_ROW = 88

    last_available_date = datetime.strptime(info['dates'][-1], '%Y-%m-%d')
    last_column = ws.max_column
    # print("last column", last_column)
    last_table_date = ws.cell(row=1, column=last_column).value
    new_days_amount = (last_available_date - last_table_date).days
    for i in range(1, new_days_amount + 1):
        ws.cell(row=1, column=(last_column + i)).value = (last_table_date + timedelta(days=i))
        ws.cell(row=1, column=(last_column + i)).number_format = 'DD.MM.YYYY'

        for row in range(*REGION_ROWS_RANGE):
            name = ws.cell(row=row, column=REGION_COLUMN).value
            key = normolize_string(name)
            if key in info['data']:
                ws.cell(row=row, column=(last_column + i)).value = \
                    info['data'][key]['cases'][(i - 1) - new_days_amount][0]
            else:
                raise KeyError(key)

        formula = ws.cell(row=OVERALL_CASES_FORMULA_ROW, column=ws.max_column - 1).value
        last_cell_id = get_column_letter(ws.max_column - 1) + str(OVERALL_CASES_FORMULA_ROW)
        next_cell_id = get_column_letter(ws.max_column) + str(OVERALL_CASES_FORMULA_ROW)
        ws[next_cell_id] = Translator(formula, origin=last_cell_id).translate_formula(next_cell_id)

    info.update({'new_days_amount': new_days_amount})


def get_regions_info() -> dict:
    final_region_data = {}

    session = HTMLSession()
    response: dict = session.get(URL, verify=False).json()
    regions = response['russia_stat_struct']['data']
    dates = response['russia_stat_struct']['dates']
    for region in regions:
        full_name = regions[region]['info']['name']
        short_name = regions[region]['info']['short_name']
        regions[region]['info'].pop('name', None)
        regions[region]['info'].pop('short_name', None)
        final_region_data.update(
            dict.fromkeys([normolize_string(full_name), normolize_string(short_name)], regions[region]))
    print(dates[-1])
    return {'data': final_region_data,
            'dates': dates,
            'last_date': datetime.strptime(dates[-1], '%Y-%m-%d')}


def continue_formula_right(ws: Worksheet, rows: list, column: int):
    for row in rows:
        formula = ws.cell(row=row, column=column - 1).value
        last_cell_id = get_column_letter(column - 1) + str(row)
        next_cell_id = get_column_letter(column) + str(row)
        ws[next_cell_id] = Translator(formula, origin=last_cell_id).translate_formula(next_cell_id)
        # print(last_cell_id, next_cell_id)
        ws[next_cell_id]._style = copy(ws[last_cell_id]._style)
        # print(ws[last_cell_id]._style, ws[next_cell_id]._style)


def continue_date_right(ws: Worksheet, rows: list):
    for row in rows:
        try:
            column = [c.value for c in ws[row]].index(None) + 1
        except ValueError:
            column = ws.max_column + 1
        last_table_date = ws.cell(row=row, column=column - 1).value
        ws.cell(row=row, column=column).value = (last_table_date) + timedelta(days=1)
        ws.cell(row=row, column=column).number_format = 'DD.MM.YYYY'

        last_cell_id = get_column_letter(column - 1) + str(row)
        next_cell_id = get_column_letter(column) + str(row)
        ws[next_cell_id]._style = copy(ws[last_cell_id]._style)


def continue_number_right(ws: Worksheet, rows: list):
    for row in rows:
        try:
            column = [c.value for c in ws[row]].index(None) + 1
        except ValueError:
            column = ws.max_column + 1
        last_num = ws.cell(row=row, column=column - 1).value
        ws.cell(row=row, column=column).value = last_num + 1

        last_cell_id = get_column_letter(column - 1) + str(row)
        next_cell_id = get_column_letter(column) + str(row)
        ws[next_cell_id]._style = copy(ws[last_cell_id]._style)


def gain_sheet(ws: Worksheet, info: dict):
    NEW_COLUMN_NUM_GAIN = [c.value for c in ws[1]].index(None) + 1

    for i in range(info['new_days_amount']):
        continue_date_right(ws, [1])
        continue_formula_right(ws, list(range(2, 96)), NEW_COLUMN_NUM_GAIN + i)

    info['last_col_in_gain_sheet'] = [c.value for c in ws[1]].index(None)


def daily_region_gain_sheet(ws: Worksheet, info: dict):
    NEW_COLUMN_NUM_DAILY_GAIN = ws.max_column + 1

    for i in range(info['new_days_amount']):
        continue_date_right(ws, [1, 20, 35])
        continue_formula_right(ws,
                               list(range(2, 19)) + list(range(21, 34)) + list(range(36, 49)),
                               NEW_COLUMN_NUM_DAILY_GAIN + i)


def weekly_gain_sheet(ws: Worksheet, info: dict):
    print(info['last_date'].weekday() + 1, WEEKLY_REPORT_DAY)
    new_col = [c.value for c in ws[1]].index(None) + 1
    ws.insert_cols(new_col)
    continue_number_right(ws, [1])
    continue_formula_right(ws, list(range(2, 94)), new_col)
    for row in range(2, 87):
        raw_formula = ws.cell(row=row, column=new_col).value
        tokenized_formula = Tokenizer(raw_formula)
        last_col_in_gain_sheet = info['last_col_in_gain_sheet']

        tokenized_formula.items[1].value = re.sub(r'(?<=:\$)\D*',
                                                  get_column_letter(last_col_in_gain_sheet),
                                                  tokenized_formula.items[1].value)
        tokenized_formula.items[3].value = re.sub(r'(?<=:\$).*(?=\$)',
                                                  get_column_letter(last_col_in_gain_sheet),
                                                  tokenized_formula.items[3].value)
        new_formula = ''.join(token.value for token in tokenized_formula.items)
        ws.cell(row=row, column=new_col).value = fr'={new_formula}'


def tpr_weekly_gain_sheet(ws: Worksheet, info: dict):
    new_col = [c.value for c in ws[1]].index(None) + 1
    ws.insert_cols(new_col)
    continue_number_right(ws, [1])
    continue_formula_right(ws, list(range(2, 88)), new_col)


# {}

# def cases_sheet(wb):
#
#
    # continue_number_right(ws_weekly_gain, )


def continue_formula_n_down(ws: Worksheet, columns: list, row: int, n: int):
    for column in columns:

        formula = ws.cell(row=(row - n), column=column).value
        # print(row-n, column, formula)
        last_cell_id = get_column_letter(column) + str(row - n)
        next_cell_id = get_column_letter(column) + str(row)
        ws[next_cell_id] = Translator(formula, origin=last_cell_id).translate_formula(next_cell_id)
        ws[next_cell_id]._style = copy(ws[last_cell_id]._style)


def continue_formula_down(ws: Worksheet, columns: list, row: int):
    continue_formula_n_down(ws, columns, row, 1)


# def get_first_clear_column_in_row(row):
#     return max((c.row for c in column if c.value is not None)) + 1


def get_first_clear_row_in_column(column):
    return max((c.row for c in column if c.value is not None)) + 1

def base_rf_sheet(ws: Worksheet, info: dict):
    NEW_ROW_NUM = get_first_clear_row_in_column(ws['B'])

    ws.cell(row=NEW_ROW_NUM, column=1).value = info['last_date']
    ws.cell(row=NEW_ROW_NUM, column=1).number_format = 'DD.MM'
    for i, stat in enumerate(info['general']):
        ws.cell(row=NEW_ROW_NUM, column=2 + i).value = info['general'][i]
        ws.cell(row=NEW_ROW_NUM, column=2 + i)._style = copy(ws.cell(row=NEW_ROW_NUM - 1, column=2 + i)._style)
    continue_formula_down(ws,
                          list(range(5, 20)) + [26] + list(range(28, 36)),
                          NEW_ROW_NUM)

    if info['last_date'].weekday() + 1 == WEEKLY_REPORT_DAY:  # if it's sunday
        continue_formula_n_down(ws, list(range(20, 25))+[27], NEW_ROW_NUM, 7)
        continue_formula_n_down(ws, list(range(20, 23)), NEW_ROW_NUM + 1, 7)


def date_week_sheet(ws: Worksheet, info: dict):
    NEW_ROW_NUM = get_first_clear_row_in_column(ws['A'])
    print(info['last_date'].weekday() + 1, WEEKLY_REPORT_DAY)
    ws.cell(row=NEW_ROW_NUM, column=1).value = ws.cell(row=NEW_ROW_NUM - 1, column=1).value + 1
    print(NEW_ROW_NUM)
    continue_formula_down(ws, list(range(2, 4)), NEW_ROW_NUM)


# def daily_gain_sheet(ws: Worksheet, actual_date: datetime):

def get_general_info(info: dict):
    STOP_COVID_URL = 'https://стопкоронавирус.рф/information/'
    session = HTMLSession()
    response = session.get(STOP_COVID_URL, verify=False)
    response.html.render()

    rows = response.html.xpath('//h3[@class="cv-stats-virus__item-value"]/text()')
    statistics = [int(''.join(row.strip('\n ').split())) for row in rows]
    useful_statistics = [statistics[2], statistics[4], statistics[0]]
    info.update({'general': useful_statistics})



def parse_info():
    info = get_regions_info()
    # {dates: [list of available dates], data: {'*REGION_NAME*': {*REGION_INFO*}}}
    get_general_info(info)
    return info


def weekly_region_gain_sheet(ws: Worksheet, info: dict):
    new_column = ws.max_column + 1

    continue_number_right(ws, [1, 20])
    continue_formula_right(ws,
                           list(range(2, 17)) + list(range(21, 34)),
                           new_column)

def gain_7day_sheet(ws: Worksheet, info: dict):
    new_column = [c.value for c in ws[1]].index(None) + 1

    for i in range(info['new_days_amount']):
        ws.insert_cols(new_column+i)
        continue_date_right(ws, [1])
        continue_formula_right(ws, list(range(2, 94)), new_column + i)


def rt_sheet(ws: Worksheet, info: dict):
    new_column = [c.value for c in ws[2]].index(None) + 1

    for i in range(info['new_days_amount']):
        ws.insert_cols(new_column+i)
        continue_date_right(ws, [2])
        continue_formula_right(ws, list(range(3, 89)), new_column + i)


def delta_day_sheet(ws: Worksheet, info: dict):
    new_column = ws.max_column + 1
    bigger_zero_fill = PatternFill(start_color='F1CCB1', end_color='F1CCB1', fill_type='solid')
    lower_zero_fill = PatternFill(start_color='E4EFDC', end_color='E4EFDC', fill_type='solid')

    for i in range(info['new_days_amount']):
        continue_date_right(ws, [1])
        ws.cell(row=2, column=new_column + i).value = DAYS_OF_THE_WEEK_MAP.get(
            ws.cell(row=1, column=new_column + i).value.weekday() + 1
        )
        last_cell_id = get_column_letter(new_column + i - 1) + str(2)
        next_cell_id = get_column_letter(new_column + i) + str(2)
        ws[next_cell_id]._style = copy(ws[last_cell_id]._style)

        continue_formula_right(ws, list(range(3, 98)), new_column + i)
        cell_ids_formatting = '{}3:{}87'.format(get_column_letter(new_column + i), get_column_letter(new_column + i))
        ws.conditional_formatting.add(cell_ids_formatting, FormulaRule(formula=[f'{next_cell_id}>0'], fill=bigger_zero_fill))
        ws.conditional_formatting.add(cell_ids_formatting, FormulaRule(formula=[f'{next_cell_id}<0'], fill=lower_zero_fill))


def msc_sp_sheet(ws: Worksheet, info: dict):
    new_column = ws.max_column + 1
    # bigger_zero_fill = PatternFill(start_color='F1CCB1', end_color='F1CCB1', fill_type='solid')
    # lower_zero_fill = PatternFill(start_color='E4EFDC', end_color='E4EFDC', fill_type='solid')

    for i in range(info['new_days_amount']):
        continue_date_right(ws, [1])
        ws.cell(row=2, column=new_column + i).value = DAYS_OF_THE_WEEK_MAP.get(
            ws.cell(row=1, column=new_column + i).value.weekday() + 1
        )
        last_cell_id = get_column_letter(new_column + i - 1) + str(2)
        next_cell_id = get_column_letter(new_column + i) + str(2)
        ws[next_cell_id]._style = copy(ws[last_cell_id]._style)
        continue_formula_right(ws, list(range(3, 98)), new_column + i)

def main(filename: str):
    wb = openpyxl.load_workbook(filename)

    info = parse_info()

    fill_сases_sheet(wb['Случаев'], info)
    gain_sheet(wb['Прирост'], info)
    msc_sp_sheet(wb['МСК и СП'], info)
    rt_sheet(wb['Rt'], info)
    gain_7day_sheet(wb['прирост 7дн'], info)
    delta_day_sheet(wb['дельта за сутки'], info)
    daily_region_gain_sheet(wb['По рег прис (сут)'], info)

    # if info['last_date'].weekday()+1 == WEEKLY_REPORT_DAY:
    #     weekly_gain_sheet(wb['Прирост нед'], info)
    #     tpr_weekly_gain_sheet(wb['Тпр нед прироста'], info)
    #     weekly_region_gain_sheet(wb['По рег прис (нед)'], info)
    #     date_week_sheet(wb['Дата-неделя'], info)

    # base_rf_sheet(wb['База РФ'], info)


    wb.save('russia_regions.xlsx')


if __name__ == '__main__':
    with open('conf_covid', 'r') as f:
        filename = f.read().strip()
    main(filename)
